"""Excel · HTML 변환 함수 (매매·전월세 공용)"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# === Excel ===

def to_excel_bytes(summary_df: pd.DataFrame, raw_df: pd.DataFrame,
                   num_cols_summary: set, num_cols_raw: set,
                   summary_widths: list[int], raw_widths: list[int]) -> bytes:
    """단지별 요약 + 전체 내역 2시트 Excel을 메모리 bytes로 반환"""
    wb = Workbook()
    dark = PatternFill("solid", fgColor="2D2D2D")
    white_bold = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
    center = Alignment(wrap_text=True, vertical="center", horizontal="center")
    left = Alignment(wrap_text=True, vertical="center", horizontal="left")
    right = Alignment(wrap_text=True, vertical="center", horizontal="right")
    thin = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    def write_sheet(ws, df, widths, num_cols):
        headers = list(df.columns)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill, cell.font, cell.alignment, cell.border = dark, white_bold, center, thin
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c, key in enumerate(headers, 1):
                v = row[c - 1]
                cell = ws.cell(r_idx, c, v)
                cell.alignment = right if key in num_cols else left
                cell.border = thin
                if key in num_cols and isinstance(v, (int, float)):
                    cell.number_format = "#,##0.00" if isinstance(v, float) else "#,##0"
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True

    ws1 = wb.active
    ws1.title = "단지별 요약"
    write_sheet(ws1, summary_df, summary_widths, num_cols_summary)

    ws2 = wb.create_sheet("전체 내역")
    write_sheet(ws2, raw_df, raw_widths, num_cols_raw)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# === HTML ===

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>{title}</title>
<style>
  body {{ font-family: "맑은 고딕", -apple-system, sans-serif; margin: 24px; color: #222; }}
  h1 {{ font-size: 20px; margin-bottom: 8px; }}
  .meta {{ color: #666; font-size: 13px; margin-bottom: 24px; }}
  .summary-cards {{ display: flex; gap: 12px; margin-bottom: 20px; flex-wrap: wrap; }}
  .card {{ background: #f6f6f6; border-radius: 8px; padding: 14px 18px; min-width: 140px; }}
  .card .label {{ font-size: 12px; color: #888; }}
  .card .value {{ font-size: 22px; font-weight: 700; color: #2D2D2D; }}
  h2 {{ font-size: 16px; margin-top: 32px; border-bottom: 2px solid #2D2D2D; padding-bottom: 6px; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 12px; margin-top: 12px; }}
  thead th {{ background: #2D2D2D; color: #fff; padding: 8px 6px; text-align: center; font-weight: 600; }}
  tbody td {{ border: 1px solid #DDD; padding: 6px; }}
  tbody tr:nth-child(even) {{ background: #fafafa; }}
  td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .footer {{ margin-top: 36px; color: #999; font-size: 11px; }}
</style>
</head>
<body>
<h1>🏢 {title}</h1>
<div class="meta">{meta}</div>

<div class="summary-cards">
{cards_html}
</div>

<h2>📋 단지별 요약</h2>
{summary_html}

<h2>📑 전체 내역</h2>
{raw_html}

<div class="footer">
  데이터 출처: 국토교통부 공공데이터포털 ({api_name}) ·
  생성: {gen_at}
</div>
</body>
</html>"""


def _df_to_html_table(df: pd.DataFrame, num_cols: set) -> str:
    cols = list(df.columns)
    th = "".join(f"<th>{c}</th>" for c in cols)
    rows_html = []
    for row in df.itertuples(index=False):
        tds = []
        for i, c in enumerate(cols):
            v = row[i]
            cls = ' class="num"' if c in num_cols else ""
            if isinstance(v, float):
                disp = f"{v:,.2f}"
            elif isinstance(v, int):
                disp = f"{v:,}"
            else:
                disp = "" if v is None else str(v)
            tds.append(f"<td{cls}>{disp}</td>")
        rows_html.append("<tr>" + "".join(tds) + "</tr>")
    return f"<table><thead><tr>{th}</tr></thead><tbody>{''.join(rows_html)}</tbody></table>"


def to_html_bytes(summary_df: pd.DataFrame, raw_df: pd.DataFrame,
                  title: str, meta: str, gen_at: str,
                  cards: list[tuple[str, str]],
                  num_cols_summary: set, num_cols_raw: set,
                  api_name: str = "RTMSDataSvcAptTrade") -> bytes:
    cards_html = "".join(
        f'<div class="card"><div class="label">{label}</div>'
        f'<div class="value">{value}</div></div>'
        for label, value in cards
    )
    html = HTML_TEMPLATE.format(
        title=title, meta=meta, cards_html=cards_html,
        summary_html=_df_to_html_table(summary_df, num_cols_summary),
        raw_html=_df_to_html_table(raw_df, num_cols_raw),
        gen_at=gen_at, api_name=api_name,
    )
    return html.encode("utf-8")
